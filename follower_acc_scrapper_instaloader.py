import instaloader
import openpyxl
import time

LIMIT = 10000

L = instaloader.Instaloader()

L.login('osakakuma', 'Okuma2020') 

wb2 = openpyxl.load_workbook('./excel/follower_list_1.xlsx')
sheet3 = wb2['Sheet1']

wb1 = openpyxl.load_workbook('./excel/follower_follower.xlsx')
sheet1 = wb1['Sheet1']
rowctr = sheet1.max_row + 2
count, curcount = 1, 0

name_col = sheet1['A']

for row in name_col:
    if row.value:
        curcount += 1
    elif row.value == None:
        continue 

print(curcount)

for row in sheet3.iter_rows():
    for cell in row:
        if curcount < count <= curcount+2:
            follower_name = cell.value
            follower_name.strip()
            print(follower_name)
            try:
                profile = instaloader.Profile.from_username(L.context, follower_name)
                sheet1['A' + str(rowctr)] = follower_name
            except:
                sheet1['A' + str(rowctr)] = follower_name
                sheet1['B' + str(rowctr)] = 'Doesnt exist'
                count += 1
                continue
            
            followerctr = 0
            for follower in profile.get_followers():
                username = follower.username
                sheet1['C' + str(followerctr+rowctr)] = username
                followerctr += 1
                time.sleep(0.5)

            sheet1['B' + str(rowctr)] = followerctr

            followeectr = 0
            for followee in profile.get_followees():
                username = followee.username
                sheet1['E' + str(followeectr+rowctr)] = username
                followeectr += 1
                time.sleep(0.5)
            
            sheet1['D' + str(rowctr)] = followeectr

            count += 1
            rowctr += 1
            if followerctr > followeectr:
                rowctr = rowctr + followerctr
            if followerctr < followeectr:
                rowctr = rowctr + followeectr

            if followerctr == 0 and followeectr == 0:
                rowctr += 1

            print(followerctr)
            print(followeectr)

        else:
            count += 1 
            continue

wb1.save('./excel/follower_follower.xlsx')
wb1.close()
wb2.close()

